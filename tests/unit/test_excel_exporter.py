# -*- coding: utf-8 -*-
"""
Unit Tests — Excel Exporter
=============================
Tester print_results_to_excel uten live Plaxis-tilkobling.
Bruker tmp_path-fixture for å skrive til en midlertidig mappe.
"""

import pytest
from pathlib import Path
from openpyxl import load_workbook

from activities.plaxis.extraction.excel_exporter import print_results_to_excel


# ---------------------------------------------------------------------------
# Testdata
# ---------------------------------------------------------------------------

FULL_RESULTS = {
    "capacity": {
        "plates": {
            "Spunt_venstre": {
                "Utgraving til bunn": {"Nx": 245.3, "Q": 89.2,  "M": 312.5},
                "FoS analyse":        {"Nx": 267.8, "Q": 95.4,  "M": 345.2},
            },
            "Spunt_høyre": {
                "Utgraving til bunn": {"Nx": 198.0, "Q": 71.0,  "M": 280.0},
            },
        },
        "node_to_node_anchors": {
            "Anker_1": {
                "Utgraving til bunn": {"N": 150.0, "Nmax": None},
            }
        },
    },
    "msf": {
        "FoS analyse": 1.32,
    },
    "displacement": {
        "plates": {
            "Spunt_venstre": {
                "Utgraving til bunn": 23.5,
                "FoS analyse":        28.1,
            }
        }
    },
}

EMPTY_RESULTS = {
    "capacity":     {},
    "msf":          {},
    "displacement": {},
}


def _job(tmp_path):
    return {"resultsPath": {"path": str(tmp_path)}}


# ---------------------------------------------------------------------------
# Filtesting
# ---------------------------------------------------------------------------

class TestExcelFileCreation:

    def test_excel_fil_blir_opprettet(self, tmp_path):
        output = print_results_to_excel(FULL_RESULTS, _job(tmp_path))
        assert Path(output).exists()

    def test_output_er_xlsx_fil(self, tmp_path):
        output = print_results_to_excel(FULL_RESULTS, _job(tmp_path))
        assert output.suffix == ".xlsx"

    def test_timestamp_i_filnavn(self, tmp_path):
        output = print_results_to_excel(FULL_RESULTS, _job(tmp_path))
        assert "Plaxis_results_" in output.name

    def test_mappe_opprettes_om_den_ikke_finnes(self, tmp_path):
        deep_path = tmp_path / "ny_mappe" / "dypere"
        output = print_results_to_excel(FULL_RESULTS, {"resultsPath": {"path": str(deep_path)}})
        assert Path(output).exists()

    def test_tom_resultatsett_krasjer_ikke(self, tmp_path):
        output = print_results_to_excel(EMPTY_RESULTS, _job(tmp_path))
        assert Path(output).exists()


# ---------------------------------------------------------------------------
# Arkstruktur
# ---------------------------------------------------------------------------

class TestExcelSheets:

    def test_summary_ark_finnes(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        assert "Summary" in wb.sheetnames

    def test_ett_ark_per_spunt(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        assert "Spunt_venstre" in wb.sheetnames
        assert "Spunt_høyre"   in wb.sheetnames

    def test_ingen_ekstra_blanke_ark(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        assert "Sheet" not in wb.sheetnames

    def test_anchor_sheet_finnes_naar_det_er_ankerdata(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        assert "Anchors_Geogrids" in wb.sheetnames

    def test_ingen_anchor_sheet_uten_ankerdata(self, tmp_path):
        resultater_uten_ankere = {**FULL_RESULTS, "capacity": {"plates": FULL_RESULTS["capacity"]["plates"]}}
        wb = load_workbook(print_results_to_excel(resultater_uten_ankere, _job(tmp_path)))
        assert "Anchors_Geogrids" not in wb.sheetnames


# ---------------------------------------------------------------------------
# Innhold
# ---------------------------------------------------------------------------

class TestExcelContents:

    def test_msf_verdi_i_summary(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        ws = wb["Summary"]
        alle_verdier = [ws.cell(r, c).value for r in range(1, ws.max_row + 1) for c in range(1, 5)]
        assert 1.32 in alle_verdier

    def test_spunt_ark_har_korrekte_kolonneoverskrifter(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        ws = wb["Spunt_venstre"]
        overskrifter = [ws.cell(1, c).value for c in range(1, 5)]
        assert "Phase"      in overskrifter
        assert "Nx (kN/m)"  in overskrifter
        assert "Q (kN/m)"   in overskrifter
        assert "M (kNm/m)"  in overskrifter

    def test_spunt_ark_har_riktige_kraftverdier(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        ws = wb["Spunt_venstre"]
        # Rad 2 og 3 er data (rad 1 er overskrifter)
        nx_verdier = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value is not None]
        assert 245.3 in nx_verdier
        assert 267.8 in nx_verdier

    def test_spunt_ark_har_maksimum_rad(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        ws = wb["Spunt_venstre"]
        alle_celler = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert "Maximum:" in alle_celler

    def test_maksimum_nx_er_korrekt(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        ws = wb["Spunt_venstre"]
        # Finn raden med "Maximum:"
        max_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Maximum:")
        assert ws.cell(max_row, 2).value == 267.8   # max(245.3, 267.8)

    def test_displacement_verdier_i_summary(self, tmp_path):
        wb = load_workbook(print_results_to_excel(FULL_RESULTS, _job(tmp_path)))
        ws = wb["Summary"]
        alle_verdier = [ws.cell(r, c).value for r in range(1, ws.max_row + 1) for c in range(1, 6)]
        assert 23.5 in alle_verdier
        assert 28.1 in alle_verdier

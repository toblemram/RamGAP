# -*- coding: utf-8 -*-
"""
Simple Excel writer for PLAXIS results - No template required
Creates a new Excel file with extraction results
"""

from openpyxl import Workbook
from pathlib import Path
from datetime import datetime


def sanitize_sheetname(name):
    """Removes invalid Excel sheet name characters"""
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in invalid:
        name = name.replace(ch, "_")
    return name[:31]


def print_results_to_excel(results, job):
    """
    Writes extracted PLAXIS results to a new Excel file.
    
    Args:
        results: Dict with capacity, msf, displacement data
        job: Job config with resultsPath
    
    Returns:
        Path to output file
    """
    
    # Ensure result folder exists
    results_folder = Path(job["resultsPath"]["path"])
    results_folder.mkdir(parents=True, exist_ok=True)
    
    # Timestamp for unique filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = results_folder / f"Plaxis_results_{timestamp}.xlsx"
    
    # Create new workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # Extract data
    capacity = results.get("capacity", {})
    msf = results.get("msf", {})
    displacement = results.get("displacement", {})
    
    # ------------------------------------------------
    # Summary Sheet
    # ------------------------------------------------
    summary = wb.create_sheet("Oppsummering", 0)
    
    row = 1
    summary.cell(row=row, column=1, value="Plaxis Resultatuttak")
    summary.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    
    # ------------------------------------------------
    # MSF Results
    # ------------------------------------------------
    if msf:
        row += 3
        summary.cell(row=row, column=1, value="MSF Resultater")
        row += 1
        summary.cell(row=row, column=1, value="Fase")
        summary.cell(row=row, column=2, value="MSF")
        
        for phase_name, value in msf.items():
            row += 1
            summary.cell(row=row, column=1, value=phase_name)
            summary.cell(row=row, column=2, value=value)
    
    # ------------------------------------------------
    # Displacement Results
    # ------------------------------------------------
    if displacement:
        row += 3
        summary.cell(row=row, column=1, value="Forskyvning (maks)")
        row += 1
        summary.cell(row=row, column=1, value="Struktur")
        summary.cell(row=row, column=2, value="Element")
        summary.cell(row=row, column=3, value="Fase")
        summary.cell(row=row, column=4, value="Verdi")
        
        for struct_type, objects in displacement.items():
            for obj_name, phases in objects.items():
                for phase_name, value in phases.items():
                    row += 1
                    summary.cell(row=row, column=1, value=struct_type)
                    summary.cell(row=row, column=2, value=obj_name)
                    summary.cell(row=row, column=3, value=phase_name)
                    summary.cell(row=row, column=4, value=value)
    
    # ------------------------------------------------
    # Capacity Results per structure
    # ------------------------------------------------
    for struct_type in ["plates", "embedded_beams"]:
        if struct_type not in capacity:
            continue
        
        for obj_name, phases in capacity[struct_type].items():
            sheet_name = sanitize_sheetname(obj_name)
            sheet = wb.create_sheet(sheet_name)
            
            # Header
            sheet.cell(row=1, column=1, value="Fase")
            sheet.cell(row=1, column=2, value="Nx (kN/m)")
            sheet.cell(row=1, column=3, value="Q (kN/m)")
            sheet.cell(row=1, column=4, value="M (kNm/m)")
            
            row = 2
            for phase_name, forces in phases.items():
                sheet.cell(row=row, column=1, value=phase_name)
                sheet.cell(row=row, column=2, value=forces.get("Nx"))
                sheet.cell(row=row, column=3, value=forces.get("Q"))
                sheet.cell(row=row, column=4, value=forces.get("M"))
                row += 1
            
            # Max values
            row += 1
            sheet.cell(row=row, column=1, value="Maksimum:")
            
            nx_values = [f.get("Nx") for f in phases.values() if f.get("Nx") is not None]
            q_values = [f.get("Q") for f in phases.values() if f.get("Q") is not None]
            m_values = [f.get("M") for f in phases.values() if f.get("M") is not None]
            
            if nx_values:
                sheet.cell(row=row, column=2, value=max(nx_values))
            if q_values:
                sheet.cell(row=row, column=3, value=max(q_values))
            if m_values:
                sheet.cell(row=row, column=4, value=max(m_values))
    
    # ------------------------------------------------
    # Anchor/Geogrid Results
    # ------------------------------------------------
    anchor_types = ["node_to_node_anchors", "fixed_end_anchors", "geogrids"]
    anchor_data = []
    
    for anchor_type in anchor_types:
        if anchor_type not in capacity:
            continue
        for obj_name, phases in capacity[anchor_type].items():
            for phase_name, forces in phases.items():
                anchor_data.append({
                    'type': anchor_type,
                    'name': obj_name,
                    'phase': phase_name,
                    'N': forces.get('N') or forces.get('Nmax')
                })
    
    if anchor_data:
        anchor_sheet = wb.create_sheet("Ankere_Geogrids")
        anchor_sheet.cell(row=1, column=1, value="Type")
        anchor_sheet.cell(row=1, column=2, value="Navn")
        anchor_sheet.cell(row=1, column=3, value="Fase")
        anchor_sheet.cell(row=1, column=4, value="N (kN)")
        
        for i, data in enumerate(anchor_data, start=2):
            anchor_sheet.cell(row=i, column=1, value=data['type'])
            anchor_sheet.cell(row=i, column=2, value=data['name'])
            anchor_sheet.cell(row=i, column=3, value=data['phase'])
            anchor_sheet.cell(row=i, column=4, value=data['N'])
    
    # Remove default sheet if it still exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Save workbook
    wb.save(output_file)
    
    print(f"Excel results written to {output_file}")
    
    return output_file

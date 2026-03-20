# -*- coding: utf-8 -*-
"""
Level 5 — Result Export to Excel
==================================
Writes extracted Plaxis results (forces, MSF, displacements) to a new
Excel file using openpyxl. No Excel template is required.

Public API:
    print_results_to_excel(results, job) -> Path
"""

from pathlib import Path
from datetime import datetime
from typing import Any, Dict

from openpyxl import Workbook


def _sanitize_sheet_name(name: str) -> str:
    """Strip characters that are illegal in Excel sheet names."""
    for ch in ('\\', '/', '*', '[', ']', ':', '?'):
        name = name.replace(ch, '_')
    return name[:31]


def print_results_to_excel(results: Dict[str, Any], job: Dict[str, Any]) -> Path:
    """
    Write extracted Plaxis results to a timestamped Excel file.

    Args:
        results: Dict with 'capacity', 'msf', and 'displacement' keys.
        job:     Job configuration dict (must contain resultsPath.path).

    Returns:
        Path to the written Excel file.
    """
    output_folder = Path(job['resultsPath']['path'])
    output_folder.mkdir(parents=True, exist_ok=True)

    timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = output_folder / f'Plaxis_results_{timestamp}.xlsx'

    wb = Workbook()

    capacity     = results.get('capacity', {})
    msf          = results.get('msf', {})
    displacement = results.get('displacement', {})

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------
    summary = wb.active
    summary.title = 'Summary'
    row = 1

    summary.cell(row=row, column=1, value='Plaxis Result Extraction')
    summary.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M'))

    # MSF
    if msf:
        row += 3
        summary.cell(row=row, column=1, value='MSF Results')
        row += 1
        summary.cell(row=row, column=1, value='Phase')
        summary.cell(row=row, column=2, value='MSF')
        for phase_name, value in msf.items():
            row += 1
            summary.cell(row=row, column=1, value=phase_name)
            summary.cell(row=row, column=2, value=value)

    # Displacement
    if displacement:
        row += 3
        summary.cell(row=row, column=1, value='Displacement (max)')
        row += 1
        for header, col in [('Structure type', 1), ('Element', 2), ('Phase', 3), ('Value', 4)]:
            summary.cell(row=row, column=col, value=header)
        for struct_type, objects in displacement.items():
            for obj_name, phases in objects.items():
                for phase_name, value in phases.items():
                    row += 1
                    summary.cell(row=row, column=1, value=struct_type)
                    summary.cell(row=row, column=2, value=obj_name)
                    summary.cell(row=row, column=3, value=phase_name)
                    summary.cell(row=row, column=4, value=value)

    # ------------------------------------------------------------------
    # One sheet per plate / embedded beam
    # ------------------------------------------------------------------
    for struct_type in ('plates', 'embedded_beams'):
        if struct_type not in capacity:
            continue
        for obj_name, phases in capacity[struct_type].items():
            sheet = wb.create_sheet(_sanitize_sheet_name(obj_name))
            sheet.cell(row=1, column=1, value='Phase')
            sheet.cell(row=1, column=2, value='Nx (kN/m)')
            sheet.cell(row=1, column=3, value='Q (kN/m)')
            sheet.cell(row=1, column=4, value='M (kNm/m)')

            data_row = 2
            for phase_name, forces in phases.items():
                sheet.cell(row=data_row, column=1, value=phase_name)
                sheet.cell(row=data_row, column=2, value=forces.get('Nx'))
                sheet.cell(row=data_row, column=3, value=forces.get('Q'))
                sheet.cell(row=data_row, column=4, value=forces.get('M'))
                data_row += 1

            # Max row
            data_row += 1
            sheet.cell(row=data_row, column=1, value='Maximum:')
            nx_vals = [f.get('Nx') for f in phases.values() if f.get('Nx') is not None]
            q_vals  = [f.get('Q')  for f in phases.values() if f.get('Q')  is not None]
            m_vals  = [f.get('M')  for f in phases.values() if f.get('M')  is not None]
            if nx_vals: sheet.cell(row=data_row, column=2, value=max(nx_vals))
            if q_vals:  sheet.cell(row=data_row, column=3, value=max(q_vals))
            if m_vals:  sheet.cell(row=data_row, column=4, value=max(m_vals))

    # ------------------------------------------------------------------
    # One sheet for anchors / geogrids
    # ------------------------------------------------------------------
    anchor_types = ('node_to_node_anchors', 'fixed_end_anchors', 'geogrids')
    anchor_rows = []
    for anchor_type in anchor_types:
        if anchor_type not in capacity:
            continue
        for obj_name, phases in capacity[anchor_type].items():
            for phase_name, forces in phases.items():
                anchor_rows.append({
                    'type':  anchor_type,
                    'name':  obj_name,
                    'phase': phase_name,
                    'N':     forces.get('N') or forces.get('Nmax'),
                })

    if anchor_rows:
        anch_sheet = wb.create_sheet('Anchors_Geogrids')
        for col, header in enumerate(('Type', 'Name', 'Phase', 'N (kN)'), start=1):
            anch_sheet.cell(row=1, column=col, value=header)
        for i, row_data in enumerate(anchor_rows, start=2):
            anch_sheet.cell(row=i, column=1, value=row_data['type'])
            anch_sheet.cell(row=i, column=2, value=row_data['name'])
            anch_sheet.cell(row=i, column=3, value=row_data['phase'])
            anch_sheet.cell(row=i, column=4, value=row_data['N'])

    # Remove the default blank sheet if a named Summary already exists
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']

    wb.save(output_file)
    print(f'Excel results written to {output_file}')
    return output_file


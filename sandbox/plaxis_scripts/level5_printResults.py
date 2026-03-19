# -*- coding: utf-8 -*-

"""
Excel writer for PLAXIS automated result extraction.

This script:
1. Opens the Excel template workbook
2. Copies the sheet "Beregningsark-spunt-Template"
3. Creates one sheet per (structure × phase)
4. Writes extracted PLAXIS forces into specific cells
5. Links selected cells back to the template sheet
6. Adds anchor/geogrid forces
7. Generates a summary sheet "Oppsummert"

The script preserves macros (.xlsm) and existing sheets.
"""

from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
from openpyxl.utils import range_boundaries


# ------------------------------------------------
# Cells that must always reference the template
# ------------------------------------------------
# These cells will not contain copied values.
# Instead they will contain formulas such as:
#
# ='Beregningsark-spunt-Template'!$F$46
#
# This allows engineers to modify parameters
# directly in the template sheet and have all
# generated sheets update automatically.
# ------------------------------------------------

LINKED_TEMPLATE_CELLS = [
    "F46",
    "F47",
    "F48",
    "F50",
    "F51",
    "F52",
    "F79",
    "F180:I183",
    "F191:I192",
    "F201:I201",
    "F208:I215",
    "F214:I215",
    "F227:I227"
]


# ------------------------------------------------
# Utility: make Excel-safe sheet names
# ------------------------------------------------
def sanitize_sheetname(name):
    """
    Removes characters Excel does not allow
    in sheet names and limits length to 31 chars.
    """

    invalid = ['\\', '/', '*', '[', ']', ':', '?']

    for ch in invalid:
        name = name.replace(ch, "_")

    return name[:31]


# ------------------------------------------------
# Create formula links back to template sheet
# ------------------------------------------------
def link_cells_to_template(sheet, template_name, cell_refs):

    """
    Writes formulas into selected cells so that
    they reference the same cells in the template sheet.

    Example result in Excel:
        ='Beregningsark-spunt-Template'!$F$46
    """

    for ref in cell_refs:

        # --- Case 1: single cell ---
        if ":" not in ref:

            col = ref[0]
            row = ref[1:]

            sheet[ref] = f"='{template_name}'!${col}${row}"

        # --- Case 2: cell range ---
        else:

            min_col, min_row, max_col, max_row = range_boundaries(ref)

            for col in range(min_col, max_col + 1):
                for row in range(min_row, max_row + 1):

                    cell = sheet.cell(row=row, column=col)

                    col_letter = cell.column_letter

                    cell.value = f"='{template_name}'!${col_letter}${row}"


# ------------------------------------------------
# Write anchor/geogrid forces
# ------------------------------------------------
def write_anchor_values(sheet, anchors):

    """
    Writes up to four anchor or geogrid forces
    into cells F39–F42.
    """

    cells = ["F39", "F40", "F41", "F42"]

    for i, val in enumerate(anchors[:4]):
        sheet[cells[i]] = val


# ------------------------------------------------
# Compute maximum forces per structure
# ------------------------------------------------
def compute_structure_maxima(capacity_results):

    """
    From the extracted capacity results, determine
    which phase gives the maximum value for Nx, Q, and M
    for each structure.

    Returns a list of dictionaries used by
    the summary sheet.
    """

    summary = []

    for struct_type, objects in capacity_results.items():

        # Anchors/geogrids handled separately
        if struct_type in [
            "fixed_end_anchors",
            "node_to_node_anchors",
            "geogrids"
        ]:
            continue

        for obj, phases in objects.items():

            # Store (value, phase)
            max_forces = {
                "Nx": (None, None),
                "Q": (None, None),
                "M": (None, None)
            }

            for phase, forces in phases.items():

                for force in ["Nx", "Q", "M"]:

                    val = forces.get(force)

                    if val is None:
                        continue

                    current_max, _ = max_forces[force]

                    if current_max is None or val > current_max:

                        max_forces[force] = (val, phase)

            summary.append({
                "structure_type": struct_type,
                "object": obj,
                "Nx": max_forces["Nx"],
                "Q": max_forces["Q"],
                "M": max_forces["M"]
            })

    return summary


# ------------------------------------------------
# Main Excel writer
# ------------------------------------------------
def print_results_to_excel(results, job):

    """
    Writes extracted PLAXIS results to Excel.

    Workflow:
        1. Open template workbook
        2. Create result sheets for each structure-phase
        3. Write forces and anchors
        4. Generate summary sheet
        5. Save workbook
    """

    base_path = Path(__file__).resolve().parent
    template_path = base_path / "excelTemplate" / "Spuntberegninger_rev0.8.xlsm"

    # Ensure result folder exists
    results_folder = Path(job["resultsPath"]["path"])
    results_folder.mkdir(parents=True, exist_ok=True)

    # Timestamp prevents overwriting previous results
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    output_file = results_folder / f"Plaxis_results_{timestamp}.xlsm"

    # Open Excel template and keep macros
    wb = load_workbook(template_path, keep_vba=True)

    template_sheet = wb["Master"]

    # Extract capacity results
    capacity = results.get("capacity", {})

    # ------------------------------------------------
    # Collect anchor/geogrid forces per phase
    # ------------------------------------------------

    anchor_forces = {}

    for anchor_type in [
        "fixed_end_anchors",
        "node_to_node_anchors",
        "geogrids"
    ]:

        if anchor_type not in capacity:
            continue

        for obj, phases in capacity[anchor_type].items():

            for phase, data in phases.items():

                val = data.get("N") or data.get("Nmax")

                anchor_forces.setdefault(phase, []).append(val)

    # ------------------------------------------------
    # Create sheets for plates and embedded beams
    # ------------------------------------------------

    for struct_type in ["plates", "embedded_beams"]:

        if struct_type not in capacity:
            continue

        for obj_name, phases in capacity[struct_type].items():

            for phase_name, forces in phases.items():

                # Build safe Excel sheet name
                sheet_name = sanitize_sheetname(
                    f"{obj_name}-{phase_name}"
                )

                # Copy the template sheet
                new_sheet = wb.copy_worksheet(template_sheet)

                new_sheet.title = sheet_name

                # Create links to template cells
                link_cells_to_template(
                    new_sheet,
                    "Master",
                    LINKED_TEMPLATE_CELLS
                )

                # Write structural forces
                new_sheet["F36"] = forces.get("M")
                new_sheet["F37"] = forces.get("Q")
                new_sheet["F38"] = forces.get("Nx")

                # Write anchor/geogrid forces
                anchors = anchor_forces.get(phase_name, [])
                write_anchor_values(new_sheet, anchors)

    # ------------------------------------------------
    # Create summary sheet
    # ------------------------------------------------

    summary_sheet = wb.create_sheet("Oppsummert")

    row = 1

    # Header row
    summary_sheet["A1"] = "Struktur"
    summary_sheet["B1"] = "Navn"
    summary_sheet["C1"] = "Nx max"
    summary_sheet["D1"] = "Nx fase"
    summary_sheet["E1"] = "Q max"
    summary_sheet["F1"] = "Q fase"
    summary_sheet["G1"] = "M max"
    summary_sheet["H1"] = "M fase"

    row = 2

    structure_summary = compute_structure_maxima(capacity)

    # Write summary values
    for item in structure_summary:

        summary_sheet.cell(row=row, column=1,
                           value=item["structure_type"])

        summary_sheet.cell(row=row, column=2,
                           value=item["object"])

        summary_sheet.cell(row=row, column=3,
                           value=item["Nx"][0])

        summary_sheet.cell(row=row, column=4,
                           value=item["Nx"][1])

        summary_sheet.cell(row=row, column=5,
                           value=item["Q"][0])

        summary_sheet.cell(row=row, column=6,
                           value=item["Q"][1])

        summary_sheet.cell(row=row, column=7,
                           value=item["M"][0])

        summary_sheet.cell(row=row, column=8,
                           value=item["M"][1])

        row += 1

    # ------------------------------------------------
    # MSF results
    # ------------------------------------------------

    msf = results.get("msf")

    if msf:

        row += 2

        summary_sheet.cell(row=row, column=1,
                           value="MSF Resultater")

        row += 1

        summary_sheet.cell(row=row, column=1,
                           value="Phase")

        summary_sheet.cell(row=row, column=2,
                           value="MSF")

        row += 1

        for phase, value in msf.items():

            summary_sheet.cell(row=row, column=1,
                               value=phase)

            summary_sheet.cell(row=row, column=2,
                               value=value)

            row += 1

    # ------------------------------------------------
    # Displacement results
    # ------------------------------------------------

    displacement = results.get("displacement")

    if displacement:

        row += 2

        summary_sheet.cell(row=row, column=1,
                           value="Forskyvning")

        row += 1

        summary_sheet.cell(row=row, column=1,
                           value="Struktur")

        summary_sheet.cell(row=row, column=2,
                           value="Navn")

        summary_sheet.cell(row=row, column=3,
                           value="Phase")

        summary_sheet.cell(row=row, column=4,
                           value="Verdi")

        row += 1

        for struct_type, objects in displacement.items():

            for obj, phases in objects.items():

                for phase, value in phases.items():

                    summary_sheet.cell(row=row, column=1,
                                       value=struct_type)

                    summary_sheet.cell(row=row, column=2,
                                       value=obj)

                    summary_sheet.cell(row=row, column=3,
                                       value=phase)

                    summary_sheet.cell(row=row, column=4,
                                       value=value)

                    row += 1

    # ------------------------------------------------
    # Save workbook
    # ------------------------------------------------

    wb.save(output_file)

    print(f"Excel results written to {output_file}")

    return output_file